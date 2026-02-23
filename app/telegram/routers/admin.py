from __future__ import annotations

import asyncio
import io
import re
from datetime import date, datetime
from typing import Any

import openpyxl
from aiogram import Router
from aiogram.enums import ChatType
from aiogram.filters import Command
from aiogram.types import Message
from pydantic import EmailStr, ValidationError

from app.core.logging import get_logger

logger = get_logger(__name__)


def _extract_command_and_args(text: str) -> tuple[str, list[str]]:
    t = (text or "").strip()
    if not t:
        return "", []
    parts = t.split()
    cmd = parts[0].lstrip("/").split("@")[0]
    return cmd, parts[1:]


def _parse_int_arg(args: list[str], idx: int = 0) -> int | None:
    if len(args) <= idx:
        return None
    try:
        return int(args[idx])
    except ValueError:
        return None


def _parse_target_user_id(message: Message) -> int | None:
    if message.reply_to_message and message.reply_to_message.from_user:
        return message.reply_to_message.from_user.id
    _cmd, args = _extract_command_and_args(message.text or "")
    return _parse_int_arg(args, 0)


def _is_admin_role(role: str) -> bool:
    return role in ("admin", "superadmin")


_HEADER_SYNONYMS: dict[str, str] = {
    "\u2116 \u043f\u0441": "ps_number",
    "\u043d\u043e\u043c\u0435\u0440 \u043f\u0441": "ps_number",
    "\u043f\u0441 \u2116": "ps_number",
    "\u043f\u0441": "ps_number",
    "ps": "ps_number",
    "ps_number": "ps_number",

    "\u043d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435 \u043f\u0441": "ps_name",
    "\u043d\u0430\u0437\u0432\u0430\u043d\u0438\u0435 \u043f\u0441": "ps_name",
    "\u043f\u0441 \u043d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435": "ps_name",
    "ps_name": "ps_name",

    "\u0432\u0438\u0434 \u0440\u0430\u0431\u043e\u0442": "work_type",
    "work_type": "work_type",

    "\u043d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435 \u043e\u0431\u044a\u0435\u043a\u0442\u0430": "title_name",
    "\u043e\u0431\u044a\u0435\u043a\u0442": "title_name",
    "title_name": "title_name",

    "\u0430\u0434\u0440\u0435\u0441": "address",
    "address": "address",

    "\u0434\u043e\u0433\u043e\u0432\u043e\u0440": "contract_number",
    "\u2116 \u0434\u043e\u0433\u043e\u0432\u043e\u0440\u0430": "contract_number",
    "\u043d\u043e\u043c\u0435\u0440 \u0434\u043e\u0433\u043e\u0432\u043e\u0440\u0430": "contract_number",
    "contract_number": "contract_number",

    "\u0437\u0430\u044f\u0432\u043a\u0430": "request_number",
    "\u2116 \u0437\u0430\u044f\u0432\u043a\u0438": "request_number",
    "\u043d\u043e\u043c\u0435\u0440 \u0437\u0430\u044f\u0432\u043a\u0438": "request_number",
    "request_number": "request_number",

    "\u0437\u0430\u043a\u0430\u0437\u0447\u0438\u043a": "customer",
    "customer": "customer",

    "\u043f\u043e\u0434\u0440\u044f\u0434\u0447\u0438\u043a": "extra.contractor",
    "contractor": "extra.contractor",
}


def _norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _cell_to_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        t = v.strip()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(t, fmt).date()
            except ValueError:
                continue
    return None


async def _download_document_bytes(message: Message) -> tuple[str, bytes] | None:
    doc = message.document
    if doc is None and message.reply_to_message is not None:
        doc = message.reply_to_message.document
    if doc is None:
        return None

    bot = message.bot
    tg_file = await bot.get_file(doc.file_id)
    buf = io.BytesIO()
    await bot.download_file(tg_file.file_path, destination=buf)
    return (doc.file_name or "objects.xlsx", buf.getvalue())


def _parse_objects_xlsx(content: bytes) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    header_row_idx: int | None = None
    headers: dict[int, str] = {}

    for r in range(1, min(10, ws.max_row) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if not any(v is not None and str(v).strip() for v in row_vals):
            continue
        header_row_idx = r
        for c, v in enumerate(row_vals, start=1):
            if v is None:
                continue
            h = _norm_header(str(v))
            if h:
                headers[c] = h
        break

    if header_row_idx is None or not headers:
        return []

    records: list[dict[str, Any]] = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        values = {c: ws.cell(row=r, column=c).value for c in headers.keys()}
        if not any(v is not None and str(v).strip() for v in values.values()):
            continue

        fields: dict[str, Any] = {}
        extra: dict[str, Any] = {}

        for c, raw_h in headers.items():
            v = values.get(c)
            if v is None:
                continue
            h = _norm_header(raw_h)
            key = _HEADER_SYNONYMS.get(h)
            if key is None:
                extra[h] = str(v).strip() if isinstance(v, str) else v
                continue

            if key.startswith("extra."):
                ex_key = key.split(".", 1)[1]
                extra[ex_key] = str(v).strip() if isinstance(v, str) else v
                continue

            if key.endswith("_start") or key.endswith("_end"):
                fields[key] = _cell_to_date(v)
                continue

            fields[key] = str(v).strip() if isinstance(v, str) else v

        if extra:
            fields["extra"] = extra

        records.append(fields)

    return records


def _dedup_key(fields: dict[str, Any]) -> str:
    ps_number = str(fields.get("ps_number") or "").strip()
    ps_name = str(fields.get("ps_name") or "").strip()
    address = str(fields.get("address") or "").strip()
    contract_number = str(fields.get("contract_number") or "").strip()

    base = "|".join([ps_number, ps_name, address, contract_number]).strip("|")
    base = re.sub(r"\s+", " ", base).lower()
    return base or (str(fields.get("title_name") or "").strip().lower() or "unknown")


def router(container: object) -> Router:  # type: ignore[type-arg]
    r = Router(name="admin")

    @r.message(Command("commands"))
    async def cmd_commands(message: Message, **kwargs: Any) -> None:
        role: str = kwargs.get("user_role", "user")
        if not _is_admin_role(role):
            await message.answer("\u26d4 \u0414\u043e\u0441\u0442\u0443\u043f \u0437\u0430\u043f\u0440\u0435\u0449\u0451\u043d.")
            return

        specs = container.registry.all_commands()  # type: ignore[attr-defined]
        # FIX BUG-2: exclude /commands itself to avoid self-reference in the list
        admin_cmds = [
            s for s in specs
            if s.required_role in ("admin", "superadmin") and s.command != "commands"
        ]
        if not admin_cmds:
            await message.answer("\U0001f4cb \u041d\u0435\u0442 \u0434\u043e\u0441\u0442\u0443\u043f\u043d\u044b\u0445 \u043a\u043e\u043c\u0430\u043d\u0434.")
            return
        lines = ["\U0001f4cb \u041a\u043e\u043c\u0430\u043d\u0434\u044b \u0430\u0434\u043c\u0438\u043d\u0438\u0441\u0442\u0440\u0430\u0442\u043e\u0440\u0430:"]
        lines.extend([f"/{s.command} \u2014 {s.description}" for s in admin_cmds])
        await message.answer("\n".join(lines))

    @r.message(Command("recipient_email"))
    async def cmd_recipient_email(message: Message, **kwargs: Any) -> None:
        role: str = kwargs.get("user_role", "user")
        if not _is_admin_role(role):
            await message.answer("\u26d4 \u0414\u043e\u0441\u0442\u0443\u043f \u0437\u0430\u043f\u0440\u0435\u0449\u0451\u043d.")
            return

        session = kwargs["session"]
        _cmd, args = _extract_command_and_args(message.text or "")
        if not args:
            cur = await container.settings_service.get_recipient_email(session)  # type: ignore[attr-defined]
            await message.answer(f"\U0001f4e7 \u0422\u0435\u043a\u0443\u0449\u0438\u0439 \u043f\u043e\u043b\u0443\u0447\u0430\u0442\u0435\u043b\u044c: {cur or '\u2014'}\n\u0424\u043e\u0440\u043c\u0430\u0442: /recipient_email user@domain")
            return

        raw = args[0].strip()
        try:
            email = EmailStr(raw)
        except (ValidationError, Exception):
            await message.answer("\u26a0\ufe0f \u041d\u0435\u043a\u043e\u0440\u0440\u0435\u043a\u0442\u043d\u044b\u0439 email. \u0424\u043e\u0440\u043c\u0430\u0442: user@domain")
            return

        await container.settings_service.set_recipient_email(session, str(email))  # type: ignore[attr-defined]
        await message.answer(f"\u2705 Email \u043f\u043e\u043b\u0443\u0447\u0430\u0442\u0435\u043b\u044f \u0443\u0441\u0442\u0430\u043d\u043e\u0432\u043b\u0435\u043d: {email}")
        logger.info("recipient_email_set", actor=message.from_user.id if message.from_user else None, email=str(email))

    @r.message(Command("time"))
    async def cmd_time(message: Message, **kwargs: Any) -> None:
        role: str = kwargs.get("user_role", "user")
        if not _is_admin_role(role):
            await message.answer("\u26d4 \u0414\u043e\u0441\u0442\u0443\u043f \u0437\u0430\u043f\u0440\u0435\u0449\u0451\u043d.")
            return

        session = kwargs["session"]
        _cmd, args = _extract_command_and_args(message.text or "")
        if not args:
            cur = await container.settings_service.get_cooldown_minutes(session)  # type: ignore[attr-defined]
            await message.answer(f"\u23f1 \u0422\u0435\u043a\u0443\u0449\u0438\u0439 cooldown: {cur} \u043c\u0438\u043d.\n\u0424\u043e\u0440\u043c\u0430\u0442: /time 30")
            return

        try:
            minutes = int(args[0])
        except ValueError:
            await message.answer("\u26a0\ufe0f \u041d\u0435\u043a\u043e\u0440\u0440\u0435\u043a\u0442\u043d\u043e\u0435 \u0447\u0438\u0441\u043b\u043e \u043c\u0438\u043d\u0443\u0442. \u0424\u043e\u0440\u043c\u0430\u0442: /time 30")
            return

        await container.settings_service.set_cooldown_minutes(session, minutes)  # type: ignore[attr-defined]
        new_val = await container.settings_service.get_cooldown_minutes(session)  # type: ignore[attr-defined]
        await message.answer(f"\u2705 Cooldown \u0443\u0441\u0442\u0430\u043d\u043e\u0432\u043b\u0435\u043d: {new_val} \u043c\u0438\u043d.")
        logger.info("cooldown_set", actor=message.from_user.id if message.from_user else None, minutes=new_val)

    @r.message(Command("object_list"))
    async def cmd_object_list(message: Message, **kwargs: Any) -> None:
        role: str = kwargs.get("user_role", "user")
        if not _is_admin_role(role):
            await message.answer("\u26d4 \u0414\u043e\u0441\u0442\u0443\u043f \u0437\u0430\u043f\u0440\u0435\u0449\u0451\u043d.")
            return

        session = kwargs["session"]
        objects = await container.objects_repo.list(session)  # type: ignore[attr-defined]
        if not objects:
            await message.answer("\u041e\u0431\u044a\u0435\u043a\u0442\u044b \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d\u044b.")
            return

        lines: list[str] = []
        for o in objects:
            title = o.ps_name or o.title_name or o.address or o.dedup_key
            lines.append(f"\u2022 {o.id} \u2014 {title}")

        await message.answer("\U0001f4cb \u041e\u0431\u044a\u0435\u043a\u0442\u044b:\n" + "\n".join(lines))

    @r.message(Command("object_add"))
    async def cmd_object_add(message: Message, **kwargs: Any) -> None:
        role: str = kwargs.get("user_role", "user")
        if not _is_admin_role(role):
            await message.answer("\u26d4 \u0414\u043e\u0441\u0442\u0443\u043f \u0437\u0430\u043f\u0440\u0435\u0449\u0451\u043d.")
            return

        text = (message.text or "")
        payload = text.split(maxsplit=1)
        if len(payload) < 2:
            await message.answer("\u0424\u043e\u0440\u043c\u0430\u0442: /object_add <ps_number>; <ps_name>; <address (\u043e\u043f\u0446.)>")
            return

        parts = [p.strip() for p in payload[1].split(";") if p.strip()]
        if len(parts) < 2:
            await message.answer("\u0424\u043e\u0440\u043c\u0430\u0442: /object_add <ps_number>; <ps_name>; <address (\u043e\u043f\u0446.)>")
            return

        ps_number, ps_name = parts[0], parts[1]
        address = parts[2] if len(parts) >= 3 else ""

        fields: dict[str, Any] = {
            "ps_number": ps_number,
            "ps_name": ps_name,
            "title_name": ps_name,
            "address": address,
            "extra": {},
        }
        dedup = _dedup_key(fields)

        session = kwargs["session"]
        obj, created = await container.objects_repo.upsert_by_dedup_key(session, dedup_key=dedup, fields=fields)  # type: ignore[attr-defined]
        await message.answer(
            f"\u2705 \u041e\u0431\u044a\u0435\u043a\u0442 {'\u0434\u043e\u0431\u0430\u0432\u043b\u0435\u043d' if created else '\u043e\u0431\u043d\u043e\u0432\u043b\u0451\u043d'}: id={obj.id}\n"
            f"\u041f\u0421: {obj.ps_number or '\u2014'} {obj.ps_name or ''}\n"
            f"\u0410\u0434\u0440\u0435\u0441: {obj.address or '\u2014'}"
        )

    @r.message(Command("object_del"))
    async def cmd_object_del(message: Message, **kwargs: Any) -> None:
        role: str = kwargs.get("user_role", "user")
        if not _is_admin_role(role):
            await message.answer("\u26d4 \u0414\u043e\u0441\u0442\u0443\u043f \u0437\u0430\u043f\u0440\u0435\u0449\u0451\u043d.")
            return

        _cmd, args = _extract_command_and_args(message.text or "")
        object_id = _parse_int_arg(args, 0)
        if object_id is None:
            await message.answer("\u0424\u043e\u0440\u043c\u0430\u0442: /object_del <object_id>")
            return

        session = kwargs["session"]
        deleted = await container.objects_repo.delete(session, object_id)  # type: ignore[attr-defined]
        if deleted:
            await message.answer(f"\u2705 \u041e\u0431\u044a\u0435\u043a\u0442 \u0443\u0434\u0430\u043b\u0451\u043d: {object_id}")
        else:
            await message.answer(f"\u2139\ufe0f \u041e\u0431\u044a\u0435\u043a\u0442 \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d: {object_id}")

    @r.message(Command("object_import"))
    async def cmd_object_import(message: Message, **kwargs: Any) -> None:
        role: str = kwargs.get("user_role", "user")
        if not _is_admin_role(role):
            await message.answer("\u26d4 \u0414\u043e\u0441\u0442\u0443\u043f \u0437\u0430\u043f\u0440\u0435\u0449\u0451\u043d.")
            return

        dl = await _download_document_bytes(message)
        if dl is None:
            await message.answer(
                "\U0001f4e5 \u041f\u0440\u0438\u0448\u043b\u0438\u0442\u0435 Excel-\u0444\u0430\u0439\u043b (.xlsx) \u0441 \u043e\u0431\u044a\u0435\u043a\u0442\u0430\u043c\u0438 \u0438 \u0432\u044b\u043f\u043e\u043b\u043d\u0438\u0442\u0435 /object_import "
                "\u0432 \u043f\u043e\u0434\u043f\u0438\u0441\u0438 \u043a \u0444\u0430\u0439\u043b\u0443 \u0438\u043b\u0438 \u043e\u0442\u0432\u0435\u0442\u043e\u043c \u043d\u0430 \u0441\u043e\u043e\u0431\u0449\u0435\u043d\u0438\u0435 \u0441 \u0444\u0430\u0439\u043b\u043e\u043c."
            )
            return

        file_name, content = dl
        await message.answer("\u23f3 \u0418\u043c\u043f\u043e\u0440\u0442\u0438\u0440\u0443\u044e \u043e\u0431\u044a\u0435\u043a\u0442\u044b \u0438\u0437 Excel...")

        records = await asyncio.to_thread(_parse_objects_xlsx, content)
        if not records:
            await message.answer("\u26a0\ufe0f \u041d\u0435 \u0443\u0434\u0430\u043b\u043e\u0441\u044c \u0440\u0430\u0441\u043f\u043e\u0437\u043d\u0430\u0442\u044c \u0442\u0430\u0431\u043b\u0438\u0446\u0443 \u0432 Excel (\u043d\u0435\u0442 \u0437\u0430\u0433\u043e\u043b\u043e\u0432\u043a\u043e\u0432/\u0441\u0442\u0440\u043e\u043a).")
            return

        session = kwargs["session"]
        created_cnt = 0
        updated_cnt = 0
        errors = 0

        for fields in records:
            try:
                dedup = _dedup_key(fields)
                _obj, was_created = await container.objects_repo.upsert_by_dedup_key(  # type: ignore[attr-defined]
                    session,
                    dedup_key=dedup,
                    fields=fields,
                )
                if was_created:
                    created_cnt += 1
                else:
                    updated_cnt += 1
            except Exception as exc:
                errors += 1
                logger.error("object_import_row_failed", file=file_name, error=str(exc)[:200])

        await message.answer(
            "\u2705 \u0418\u043c\u043f\u043e\u0440\u0442 \u0437\u0430\u0432\u0435\u0440\u0448\u0451\u043d.\n"
            f"\u0424\u0430\u0439\u043b: {file_name}\n"
            f"\u0421\u0442\u0440\u043e\u043a \u043e\u0431\u0440\u0430\u0431\u043e\u0442\u0430\u043d\u043e: {len(records)}\n"
            f"\u0421\u043e\u0437\u0434\u0430\u043d\u043e: {created_cnt}\n"
            f"\u041e\u0431\u043d\u043e\u0432\u043b\u0435\u043d\u043e: {updated_cnt}\n"
            f"\u041e\u0448\u0438\u0431\u043e\u043a: {errors}"
        )

    @r.message(Command("group_list"))
    async def cmd_group_list(message: Message, **kwargs: Any) -> None:
        role: str = kwargs.get("user_role", "user")
        if not _is_admin_role(role):
            await message.answer("\u26d4 \u0414\u043e\u0441\u0442\u0443\u043f \u0437\u0430\u043f\u0440\u0435\u0449\u0451\u043d.")
            return

        session = kwargs["session"]
        chat_id = message.chat.id if message.chat.type in (ChatType.GROUP, ChatType.SUPERGROUP) else None
        links = await container.objects_repo.list_group_links(session, chat_id=chat_id)  # type: ignore[attr-defined]
        if not links:
            await message.answer("\u041f\u0440\u0438\u0432\u044f\u0437\u043a\u0438 \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d\u044b.")
            return

        text_lines = ["\U0001f4cb \u041f\u0440\u0438\u0432\u044f\u0437\u043a\u0438 (object_id \u2192 chat_id):"]
        text_lines.extend([f"\u2022 {obj_id} \u2192 {gid}" for obj_id, gid in links])
        await message.answer("\n".join(text_lines))

    @r.message(Command("group_add"))
    async def cmd_group_add(message: Message, **kwargs: Any) -> None:
        role: str = kwargs.get("user_role", "user")
        if not _is_admin_role(role):
            await message.answer("\u26d4 \u0414\u043e\u0441\u0442\u0443\u043f \u0437\u0430\u043f\u0440\u0435\u0449\u0451\u043d.")
            return
        if message.chat.type not in (ChatType.GROUP, ChatType.SUPERGROUP):
            await message.answer("\u26a0\ufe0f /group_add \u0434\u043e\u0441\u0442\u0443\u043f\u043d\u0430 \u0442\u043e\u043b\u044c\u043a\u043e \u0432 \u0433\u0440\u0443\u043f\u043f\u0435/\u0441\u0443\u043f\u0435\u0440\u0433\u0440\u0443\u043f\u043f\u0435.")
            return

        _cmd, args = _extract_command_and_args(message.text or "")
        object_id = _parse_int_arg(args, 0)
        if object_id is None:
            await message.answer("\u0424\u043e\u0440\u043c\u0430\u0442: /group_add <object_id> (\u0432 \u0433\u0440\u0443\u043f\u043f\u0435).")
            return

        session = kwargs["session"]
        obj = await container.objects_repo.get_by_id(session, object_id)  # type: ignore[attr-defined]
        if not obj:
            await message.answer(f"\u26a0\ufe0f \u041e\u0431\u044a\u0435\u043a\u0442 \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d: {object_id}")
            return

        await container.objects_repo.link_group(session, object_id=object_id, chat_id=message.chat.id)  # type: ignore[attr-defined]
        await message.answer(f"\u2705 \u0413\u0440\u0443\u043f\u043f\u0430 \u043f\u0440\u0438\u0432\u044f\u0437\u0430\u043d\u0430 \u043a \u043e\u0431\u044a\u0435\u043a\u0442\u0443 {object_id} (chat_id={message.chat.id})")

    @r.message(Command("group_del"))
    async def cmd_group_del(message: Message, **kwargs: Any) -> None:
        role: str = kwargs.get("user_role", "user")
        if not _is_admin_role(role):
            await message.answer("\u26d4 \u0414\u043e\u0441\u0442\u0443\u043f \u0437\u0430\u043f\u0440\u0435\u0449\u0451\u043d.")
            return
        if message.chat.type not in (ChatType.GROUP, ChatType.SUPERGROUP):
            await message.answer("\u26a0\ufe0f /group_del \u0434\u043e\u0441\u0442\u0443\u043f\u043d\u0430 \u0442\u043e\u043b\u044c\u043a\u043e \u0432 \u0433\u0440\u0443\u043f\u043f\u0435/\u0441\u0443\u043f\u0435\u0440\u0433\u0440\u0443\u043f\u043f\u0435.")
            return

        _cmd, args = _extract_command_and_args(message.text or "")
        object_id = _parse_int_arg(args, 0)
        if object_id is None:
            await message.answer("\u0424\u043e\u0440\u043c\u0430\u0442: /group_del <object_id> (\u0432 \u0433\u0440\u0443\u043f\u043f\u0435).")
            return

        session = kwargs["session"]
        removed = await container.objects_repo.unlink_group(session, object_id=object_id, chat_id=message.chat.id)  # type: ignore[attr-defined]
        if removed:
            await message.answer(f"\u2705 \u041f\u0440\u0438\u0432\u044f\u0437\u043a\u0430 \u0443\u0434\u0430\u043b\u0435\u043d\u0430: object_id={object_id} \u2194 chat_id={message.chat.id}")
        else:
            await message.answer("\u2139\ufe0f \u041f\u0440\u0438\u0432\u044f\u0437\u043a\u0430 \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d\u0430.")

    @r.message(Command("user_list"))
    async def cmd_user_list(message: Message, **kwargs: Any) -> None:
        role: str = kwargs.get("user_role", "user")
        if not _is_admin_role(role):
            await message.answer("\u26d4 \u0414\u043e\u0441\u0442\u0443\u043f \u0437\u0430\u043f\u0440\u0435\u0449\u0451\u043d.")
            return

        session = kwargs["session"]
        users = await container.users_repo.list_allowed_private(session)  # type: ignore[attr-defined]
        if not users:
            await message.answer("\u0421\u043f\u0438\u0441\u043e\u043a \u043f\u0443\u0441\u0442.")
            return
        await message.answer(
            "\U0001f464 \u0420\u0430\u0437\u0440\u0435\u0448\u0451\u043d\u043d\u044b\u0435 (\u043b\u0438\u0447\u043a\u0430):\n"
            + "\n".join(f"\u2022 {u.telegram_user_id}" for u in users)
        )

    @r.message(Command("user_add"))
    async def cmd_user_add(message: Message, **kwargs: Any) -> None:
        role: str = kwargs.get("user_role", "user")
        if not _is_admin_role(role):
            await message.answer("\u26d4 \u0414\u043e\u0441\u0442\u0443\u043f \u0437\u0430\u043f\u0440\u0435\u0449\u0451\u043d.")
            return

        target_id = _parse_target_user_id(message)
        if target_id is None:
            await message.answer("\u0424\u043e\u0440\u043c\u0430\u0442: /user_add <telegram_user_id> (\u0438\u043b\u0438 \u043e\u0442\u0432\u0435\u0442\u044c\u0442\u0435 \u043d\u0430 \u0441\u043e\u043e\u0431\u0449\u0435\u043d\u0438\u0435 \u043f\u043e\u043b\u044c\u0437\u043e\u0432\u0430\u0442\u0435\u043b\u044f).")
            return

        session = kwargs["session"]
        await container.users_repo.set_allowed_private(session, target_id, True)  # type: ignore[attr-defined]
        await message.answer(f"\u2705 \u041f\u043e\u043b\u044c\u0437\u043e\u0432\u0430\u0442\u0435\u043b\u044c \u0440\u0430\u0437\u0440\u0435\u0448\u0451\u043d \u0432 \u043b\u0438\u0447\u043d\u043e\u043c \u0447\u0430\u0442\u0435: {target_id}")

    @r.message(Command("user_del"))
    async def cmd_user_del(message: Message, **kwargs: Any) -> None:
        role: str = kwargs.get("user_role", "user")
        if not _is_admin_role(role):
            await message.answer("\u26d4 \u0414\u043e\u0441\u0442\u0443\u043f \u0437\u0430\u043f\u0440\u0435\u0449\u0451\u043d.")
            return

        target_id = _parse_target_user_id(message)
        if target_id is None:
            await message.answer("\u0424\u043e\u0440\u043c\u0430\u0442: /user_del <telegram_user_id> (\u0438\u043b\u0438 \u043e\u0442\u0432\u0435\u0442\u044c\u0442\u0435 \u043d\u0430 \u0441\u043e\u043e\u0431\u0449\u0435\u043d\u0438\u0435 \u043f\u043e\u043b\u044c\u0437\u043e\u0432\u0430\u0442\u0435\u043b\u044f).")
            return

        session = kwargs["session"]
        await container.users_repo.set_allowed_private(session, target_id, False)  # type: ignore[attr-defined]
        await message.answer(f"\u2705 \u041f\u043e\u043b\u044c\u0437\u043e\u0432\u0430\u0442\u0435\u043b\u044c \u0437\u0430\u043f\u0440\u0435\u0449\u0451\u043d \u0432 \u043b\u0438\u0447\u043d\u043e\u043c \u0447\u0430\u0442\u0435: {target_id}")

    return r
