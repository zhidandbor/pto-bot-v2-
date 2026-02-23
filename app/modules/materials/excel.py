"""Excel generation for the Materials module.

This module fills a pre-defined XLSX template (template_materials.xlsx) with
object metadata and parsed material lines.

The layout is intentionally tied to the template (fixed cell addresses and
row ranges) and must be updated together with the template when it changes.
"""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import MergedCell

from app.core.logging import get_logger
from app.modules.materials.schemas import MaterialDraft

logger = get_logger(__name__)

TEMPLATE_PATH = Path(__file__).parent / "template_materials.xlsx"

ITEMS_START_ROW = 12   # первая строка позиций (B12:F36)
ITEMS_END_ROW   = 36   # последняя строка позиций
MAX_LINES       = 25   # FRMAT15

# Индексы колонок (1-based, openpyxl)
COL_A = 1   # п/п        (нумерация)
COL_B = 2   # наименование
COL_C = 3   # тип / марка
COL_E = 5   # единица измерения
COL_F = 6   # количество

_MONTHS_RU = (
    "января", "февраля", "марта", "апреля", "мая", "июня",
    "июля", "августа", "сентября", "октября", "ноября", "декабря",
)

_EXCEL_DANGEROUS_PREFIXES = ("=", "+", "-", "@")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def fill_excel_template(draft: MaterialDraft, object_data: dict[str, Any]) -> bytes:
    """Fill template_materials.xlsx with draft data and return XLSX bytes.

    object_data is expected to contain resolved object fields:
        ps_name, contractor, work_type, contract_number,
        work_period, customer, address.

    Raises:
        FileNotFoundError: template file is missing.
        ValueError: draft.lines is empty.
    """
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Шаблон не найден: {TEMPLATE_PATH}")

    if not draft.lines:
        raise ValueError("Список позиций пуст — нечего записывать в Excel")

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # --- Блок C1-C7: данные объекта (FRMAT12) ---
    _set(ws, "C1", object_data.get("ps_name", ""))
    _set(ws, "C2", object_data.get("contractor", ""))
    _set(ws, "C3", object_data.get("work_type", ""))
    _set(ws, "C4", object_data.get("contract_number", ""))
    _set(ws, "C5", object_data.get("work_period", ""))
    _set(ws, "C6", object_data.get("customer", ""))
    _set(ws, "C7", object_data.get("address", ""))

    # --- Поля заявки (FRMAT14) ---
    _set(ws, "H9",  draft.request_number)
    _set(ws, "H10", draft.user_full_name or "")
    _set(ws, "B39", f"г. Санкт-Петербург, {_ru_date(draft.request_date)}")

    # --- Очистка строк позиций (FRMAT15) ---
    _clear_items(ws)

    # --- Заполнение строк B12:F36 (FRMAT15) ---
    for line in draft.lines:
        if line.line_no > MAX_LINES:
            logger.warning(
                "excel_line_overflow",
                draft_id=draft.draft_id,
                line_no=line.line_no,
                max=MAX_LINES,
            )
            break
        row = ITEMS_START_ROW + line.line_no - 1
        _set_col(ws, row, COL_A, line.line_no)
        _set_col(ws, row, COL_B, line.name)
        _set_col(ws, row, COL_C, line.type_mark or "")
        _set_col(ws, row, COL_E, line.unit)
        _set_col(ws, row, COL_F, _format_qty(line.qty))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    logger.info(
        "excel_generated",
        draft_id=draft.draft_id,
        lines=len(draft.lines),
        ps_number=draft.ps_number,
    )
    return buf.read()


def build_file_name(draft: MaterialDraft) -> str:
    """Build a stable XLSX filename for the generated request."""
    ps = (draft.ps_number or "объект").replace(" ", "_").replace("/", "-")
    d = draft.request_date.strftime("%Y-%m-%d")
    return f"Заявка_{ps}_{d}_№{draft.counter}.xlsx"


# ---------------------------------------------------------------------------
# Security helpers
# ---------------------------------------------------------------------------

def sanitize_excel_text(value: str) -> str:
    """Prevent Excel formula injection for user-provided strings.

    If the string starts with a dangerous prefix (=, +, -, @), it is prefixed
    with an apostrophe.
    """
    stripped = value.lstrip()
    if stripped and stripped[0] in _EXCEL_DANGEROUS_PREFIXES:
        logger.warning(
            "excel_formula_injection_sanitized",
            prefix=stripped[0],
            value=value[:80],
        )
        return "'" + value
    return value


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

def _clear_items(ws: Any) -> None:
    """Clear item cells in the template (A/F columns, ITEMS_START_ROW..ITEMS_END_ROW)."""
    for row in range(ITEMS_START_ROW, ITEMS_END_ROW + 1):
        for col in (COL_A, COL_B, COL_C, COL_E, COL_F):
            _set_col(ws, row, col, None)


def _set(ws: Any, cell_ref: str, value: Any) -> None:
    """Set cell value by A1 reference.

    Skips child cells of merged regions and writes into the merge master instead.
    All string values pass through sanitize_excel_text().
    """
    if isinstance(value, str):
        value = sanitize_excel_text(value)

    cell = ws[cell_ref]
    if not isinstance(cell, MergedCell):
        cell.value = value
    else:
        master = _find_merge_master(ws, cell_ref)
        if master:
            master.value = value
        else:
            logger.warning("excel_merged_cell_skip", ref=cell_ref, value=str(value)[:40])


def _set_col(ws: Any, row: int, col: int, value: Any) -> None:
    """Set cell value by row/column (1-based)."""
    if isinstance(value, str):
        value = sanitize_excel_text(value)

    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def _find_merge_master(ws: Any, cell_ref: str) -> Any | None:
    """Find a merge-master cell for a given A1 reference."""
    from openpyxl.utils import coordinate_to_tuple, get_column_letter

    try:
        row, col = coordinate_to_tuple(cell_ref)
    except Exception:
        return None

    for merge_range in ws.merged_cells.ranges:
        if (merge_range.min_row <= row <= merge_range.max_row and
                merge_range.min_col <= col <= merge_range.max_col):
            master_ref = (
                f"{get_column_letter(merge_range.min_col)}{merge_range.min_row}"
            )
            master = ws[master_ref]
            return master if not isinstance(master, MergedCell) else None
    return None


def _ru_date(d: date) -> str:
    """Format date in Russian: '21 февраля 2026 г.'"""
    return f"{d.day} {_MONTHS_RU[d.month - 1]} {d.year} г."


def _format_qty(qty: Decimal) -> int | float:
    """Convert Decimal qty for Excel output.

    - Whole numbers are written as int.
    - Fractional values are quantized to 0.001 and written as float.
    """
    if qty == qty.to_integral_value():
        return int(qty)
    return float(qty.quantize(Decimal("0.001")))
