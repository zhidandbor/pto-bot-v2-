from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

import openpyxl

from app.core.logging import get_logger

logger = get_logger(__name__)

_SHEET_NAME = "objects"
_DATA_START_ROW = 3  # строка 1 — заголовок листа, строка 2 — заголовки столбцов


def _str(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s or None


def _date(val: Any, row_num: int, col: str) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%y", "%m/%d/%Y", "%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    logger.warning("date_parse_failed", row=row_num, col=col, value=s)
    return None


def _int(val: Any) -> int | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    return int(s) if s else None


@dataclass(frozen=True)
class ObjectRow:
    """Разобранная строка из data_objects.xlsx (лист objects, данные с 3-й строки)."""

    # --- Col A–M: прямые поля Object ---
    chat_id: int | None            # A
    ps_number: str | None          # B
    ps_name: str | None            # C
    work_type: str | None          # D
    title_name: str | None         # E
    address: str | None            # F
    contract_number: str | None    # G
    contract_start: date | None    # H
    contract_end: date | None      # I
    request_number: str | None     # J
    work_start: date | None        # K
    work_end: date | None          # L
    customer: str | None           # M

    # --- Col N–X: контакты → extra JSONB ---
    customer_position_1: str | None         # N
    customer_fio_1: str | None              # O
    customer_name_patronymic_1: str | None  # P
    customer_position_2: str | None         # Q
    customer_fio_2: str | None              # R
    customer_name_patronymic_2: str | None  # S
    contractor: str | None                  # T
    contractor_oks_name: str | None         # U
    contractor_oks_tel: str | None          # V
    contractor_osovn_name: str | None       # W
    contractor_osovn_tel: str | None        # X

    row_num: int = 0

    @property
    def dedup_key(self) -> str:
        """Уникальный ключ: № ПС | Вид работ."""
        return f"{(self.ps_number or '').strip()}|{(self.work_type or '').strip()}"

    @property
    def is_valid(self) -> bool:
        """Обязательны ps_number и work_type для формирования dedup_key."""
        return bool(self.ps_number and self.work_type)


@dataclass
class ExcelReadResult:
    rows: list[ObjectRow] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


class ObjectsExcelReader:
    """Читает data_objects.xlsx.

    Лист: 'objects'.
    Строка 2 — заголовки (игнорируется).
    Данные с строки 3.
    Новые столбцы правее X попадают в extra через ExcelImportService.
    """

    def read(
        self, file_bytes: bytes, *, filename: str = "data_objects.xlsx"
    ) -> ExcelReadResult:
        result = ExcelReadResult()
        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(file_bytes), read_only=True, data_only=True
            )
        except Exception as exc:
            result.errors.append(f"Не удалось открыть файл '{filename}': {exc}")
            logger.error("excel_open_failed", filename=filename, error=str(exc))
            return result

        if _SHEET_NAME not in wb.sheetnames:
            result.errors.append(
                f"Лист '{_SHEET_NAME}' не найден в файле '{filename}'. "
                f"Доступные листы: {wb.sheetnames}"
            )
            return result

        ws = wb[_SHEET_NAME]
        for row in ws.iter_rows(min_row=_DATA_START_ROW):
            row_num: int = row[0].row  # type: ignore[union-attr]
            cells = [c.value for c in row]

            # Пропускаем полностью пустые строки
            if all(c is None for c in cells[:24]):
                continue

            def c(idx: int) -> Any:
                return cells[idx] if idx < len(cells) else None

            try:
                obj_row = ObjectRow(
                    chat_id=_int(c(0)),
                    ps_number=_str(c(1)),
                    ps_name=_str(c(2)),
                    work_type=_str(c(3)),
                    title_name=_str(c(4)),
                    address=_str(c(5)),
                    contract_number=_str(c(6)),
                    contract_start=_date(c(7), row_num, "H"),
                    contract_end=_date(c(8), row_num, "I"),
                    request_number=_str(c(9)),
                    work_start=_date(c(10), row_num, "K"),
                    work_end=_date(c(11), row_num, "L"),
                    customer=_str(c(12)),
                    customer_position_1=_str(c(13)),
                    customer_fio_1=_str(c(14)),
                    customer_name_patronymic_1=_str(c(15)),
                    customer_position_2=_str(c(16)),
                    customer_fio_2=_str(c(17)),
                    customer_name_patronymic_2=_str(c(18)),
                    contractor=_str(c(19)),
                    contractor_oks_name=_str(c(20)),
                    contractor_oks_tel=_str(c(21)),
                    contractor_osovn_name=_str(c(22)),
                    contractor_osovn_tel=_str(c(23)),
                    row_num=row_num,
                )
            except Exception as exc:
                result.errors.append(f"Строка {row_num}: ошибка парсинга — {exc}")
                logger.error("excel_row_parse_error", row=row_num, error=str(exc))
                continue

            if not obj_row.is_valid:
                result.errors.append(
                    f"Строка {row_num}: пропущена — отсутствует № ПС или Вид работ "
                    f"(ps={obj_row.ps_number!r}, work_type={obj_row.work_type!r})"
                )
                continue

            result.rows.append(obj_row)

        logger.info(
            "excel_read_done",
            filename=filename,
            rows=len(result.rows),
            skipped_errors=len(result.errors),
        )
        return result
